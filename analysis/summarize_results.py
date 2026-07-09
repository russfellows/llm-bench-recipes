#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""
Summarize a benchmark sweep's results directory (or its .tgz archive) into
a single Excel workbook.

Reads the standard llm-bench-recipes results layout (see recipes/README.md):

    <results_dir>/
        provenance.json
        summary.csv
        <model>_<variant>_tp<TP>_isl<ISL>_osl<OSL>_c<CONC>.json
        server_tp<TP>*.log

and produces a workbook with:
    Results     - one row per (TP, ISL, OSL, CONC) combo, every field from
                  the bench-client JSON (vllm bench serve / atom output).
                  With --split-by-tp, this becomes one sheet per TP value
                  ("TP=1", "TP=2", ...) instead of a single combined sheet.
    Run Info    - provenance.json metadata (model, vendor, image, GPUs, git
                  commit, sweep matrix).
    Recipe TOML - the recipe.toml snapshot captured in provenance.json.

Usage:
    uv run analysis/summarize_results.py <results.tgz | results_dir> [-o out.xlsx] [--split-by-tp]

If -o/--output is omitted, the workbook is named
"<model>_<variant>_<timestamp>_summary.xlsx" (or "..._summary_by_tp.xlsx"
with --split-by-tp) and written next to the input.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import tarfile
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Canonical field order for bench-client result JSON (vllm bench serve /
# atom.benchmarks.benchmark_serving). Verified identical across every combo
# in a sweep at the time this was written. Any keys not in this list are
# discovered at parse time and appended to the right of the sheet rather
# than silently dropped, so a future stack adding fields doesn't lose data.
RESULT_FIELD_ORDER = [
    "date", "endpoint_type", "backend", "label", "model_id", "tokenizer_id",
    "num_prompts", "request_rate", "burstiness", "max_concurrency",
    "duration", "completed", "failed",
    "total_input_tokens", "total_output_tokens",
    "request_throughput", "request_goodput", "output_throughput",
    "total_token_throughput", "max_output_tokens_per_s",
    "max_concurrent_requests", "rtfx",
    "mean_ttft_ms", "median_ttft_ms", "std_ttft_ms",
    "p25_ttft_ms", "p50_ttft_ms", "p75_ttft_ms", "p90_ttft_ms", "p95_ttft_ms", "p99_ttft_ms",
    "mean_tpot_ms", "median_tpot_ms", "std_tpot_ms",
    "p25_tpot_ms", "p50_tpot_ms", "p75_tpot_ms", "p90_tpot_ms", "p95_tpot_ms", "p99_tpot_ms",
    "mean_itl_ms", "median_itl_ms", "std_itl_ms",
    "p25_itl_ms", "p50_itl_ms", "p75_itl_ms", "p90_itl_ms", "p95_itl_ms", "p99_itl_ms",
    "mean_e2el_ms", "median_e2el_ms", "std_e2el_ms",
    "p25_e2el_ms", "p50_e2el_ms", "p75_e2el_ms", "p90_e2el_ms", "p95_e2el_ms", "p99_e2el_ms",
]

# Columns from this set are left as text (not given a numeric display mask)
# even though some rows may coincidentally parse as numbers.
TEXT_FIELDS = {
    "date", "endpoint_type", "backend", "label",
    "model_id", "tokenizer_id", "request_rate",
}

SUMMARY_FIELDS = ["tp", "isl", "osl", "conc", "status", "result_file"]


def find_results_dir(root: Path) -> Path:
    """Locate the timestamped results dir (the one holding summary.csv)
    under `root`. `root` may already *be* that directory."""
    if (root / "summary.csv").is_file():
        return root
    matches = sorted(root.rglob("summary.csv"))
    if not matches:
        sys.exit(f"No summary.csv found under {root}")
    if len(matches) > 1:
        listing = "\n".join(f"  {m.parent}" for m in matches)
        sys.exit(
            f"Multiple summary.csv files found under {root} - point at one "
            f"results directory directly:\n{listing}"
        )
    return matches[0].parent


def extract_if_archive(input_path: Path, tmp_dir: Path) -> Path:
    if input_path.is_dir():
        return input_path
    if tarfile.is_tarfile(input_path):
        with tarfile.open(input_path) as tf:
            try:
                tf.extractall(tmp_dir, filter="data")
            except TypeError:
                # Python < 3.12 (without the backported `filter` kwarg).
                tf.extractall(tmp_dir)
        return tmp_dir
    sys.exit(f"{input_path} is neither a directory nor a tar archive")


def load_summary_rows(results_dir: Path) -> list[dict]:
    with (results_dir / "summary.csv").open(newline="") as f:
        return list(csv.DictReader(f))


def _flatten_nested_dicts(data: dict) -> dict:
    """Flatten one level of dict-valued fields into separate scalar
    columns. openpyxl cells can only hold scalars (str/int/float/bool/
    None/datetime); vLLM's bench-serve output is already flat
    (p25_ttft_ms, ...), but bench_openai.py's stack-agnostic client nests
    percentiles instead (percentiles_ttft_ms: {"25": ..., "50": ...}).
    Handled generically here rather than special-cased to one client's
    schema, so any future bench tool that nests a field doesn't crash the
    same way."""
    flat = {}
    for k, v in data.items():
        if isinstance(v, dict):
            for subk, subv in v.items():
                flat[f"{k}_p{subk}"] = subv
        else:
            flat[k] = v
    return flat


def load_result_json(results_dir: Path, row: dict) -> dict | None:
    if row.get("status") != "ok":
        return None
    result_file = row.get("result_file", "")
    path = results_dir / result_file
    if not path.is_file():
        return None
    with path.open() as f:
        data = json.load(f)
    return _flatten_nested_dicts(data)


def numeric_sort_key(row: dict):
    def num(key):
        try:
            return int(row[key])
        except (KeyError, ValueError, TypeError):
            return 0
    return (num("tp"), num("isl"), num("osl"), num("conc"))


def group_rows_by_tp(rows: list[dict]) -> dict[int, list[dict]]:
    """Group rows by integer TP value, sorted ascending. Unparseable TP
    values (shouldn't happen given summary.csv's schema) fall into 0."""
    groups: dict[int, list[dict]] = {}
    for row in rows:
        try:
            tp = int(row.get("tp", 0))
        except (TypeError, ValueError):
            tp = 0
        groups.setdefault(tp, []).append(row)
    return dict(sorted(groups.items()))


def build_results_sheet(
    ws: Worksheet, results_dir: Path, rows: list[dict], include_tp_column: bool = True,
) -> None:
    summary_fields = SUMMARY_FIELDS if include_tp_column else [f for f in SUMMARY_FIELDS if f != "tp"]

    extra_keys: list[str] = []
    seen_extra: set[str] = set()
    parsed_rows: list[dict | None] = []
    for row in rows:
        data = load_result_json(results_dir, row)
        parsed_rows.append(data)
        if data:
            for k in data:
                if k not in RESULT_FIELD_ORDER and k not in seen_extra:
                    seen_extra.add(k)
                    extra_keys.append(k)

    header = summary_fields + RESULT_FIELD_ORDER + extra_keys
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    order = sorted(range(len(rows)), key=lambda i: numeric_sort_key(rows[i]))
    for i in order:
        row, data = rows[i], parsed_rows[i]
        out = [row.get(f, "") for f in summary_fields]
        for f in RESULT_FIELD_ORDER + extra_keys:
            out.append(data.get(f, "") if data else "")
        ws.append(out)

    ws.auto_filter.ref = ws.dimensions
    for col_idx, name in enumerate(header, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(10, min(22, len(name) + 2))

    # Round floating-point metric columns for on-screen readability.
    # number_format is a display mask only - underlying precision is kept.
    numeric_col_idxs = [
        i for i, h in enumerate(header, start=1)
        if h not in SUMMARY_FIELDS and h not in TEXT_FIELDS
    ]
    for row_cells in ws.iter_rows(min_row=2):
        for col_idx in numeric_col_idxs:
            cell = row_cells[col_idx - 1]
            if isinstance(cell.value, float):
                cell.number_format = "0.000"


def build_run_info_sheet(ws: Worksheet, prov: dict) -> None:
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def add(label, value):
        ws.append([label, "" if value is None else value])

    add("timestamp_utc", prov.get("timestamp_utc"))
    add("timestamp_local", prov.get("timestamp_local"))
    add("model_name", prov.get("model_name"))
    add("variant_name", prov.get("variant_name"))
    add("model_id", prov.get("model_id"))
    add("vendor", prov.get("vendor"))
    add("stack", prov.get("stack"))

    image = prov.get("image") or {}
    add("image.ref", image.get("ref"))
    info = image.get("info") or {}
    add("image.id", info.get("id"))
    add("image.repo_digests", ", ".join(info.get("repo_digests", []) or []))

    build = prov.get("build")
    if build:
        add("build.base_image", build.get("base_image"))
        add("build.dockerfile_path", build.get("dockerfile_path"))
        add("build.build_args", json.dumps(build.get("build_args", {})))

    sweep = prov.get("sweep") or {}
    add("sweep.tp", ", ".join(sweep.get("tp", []) or []))
    add("sweep.isl_osl", ", ".join(sweep.get("isl_osl", []) or []))
    add("sweep.conc", ", ".join(sweep.get("conc", []) or []))

    system = prov.get("system") or {}
    add("system.hostname", system.get("hostname"))
    add("system.kernel", system.get("kernel"))
    add("system.os_release", " | ".join(system.get("os_release", []) or []))
    gpus = system.get("gpus") or {}
    add("system.gpus.nvidia", "\n".join(gpus.get("nvidia") or []))
    add("system.gpus.amd", "\n".join(gpus.get("amd") or []) if gpus.get("amd") else "")
    add("system.gpus.lspci", "\n".join(gpus.get("lspci") or []))

    repo = prov.get("llm_bench_recipes_repo") or {}
    add("llm_bench_recipes_repo.path", repo.get("path"))
    add("llm_bench_recipes_repo.commit", repo.get("commit"))
    add("llm_bench_recipes_repo.short", repo.get("short"))
    add("llm_bench_recipes_repo.dirty", repo.get("dirty"))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    for row_cells in ws.iter_rows(min_row=2):
        row_cells[1].alignment = Alignment(wrap_text=True, vertical="top")


def build_recipe_sheet(ws: Worksheet, prov: dict) -> None:
    recipe_toml = prov.get("recipe_toml")
    ws.append(["recipe.toml (as captured at run time)"])
    ws["A1"].font = Font(bold=True)
    if recipe_toml:
        for i, line in enumerate(recipe_toml.splitlines(), start=2):
            cell = ws.cell(row=i, column=1, value=line)
            cell.font = Font(name="Courier New", size=10)
    ws.column_dimensions["A"].width = 100


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", type=Path, help="results .tgz/.tar.gz or a results directory")
    ap.add_argument(
        "-o", "--output", type=Path, default=None,
        help="output .xlsx path (default: derived from provenance, written next to input)",
    )
    ap.add_argument(
        "--split-by-tp", action="store_true",
        help="one sheet per TP value ('TP=1', 'TP=2', ...) instead of a single combined Results sheet",
    )
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"Input not found: {args.input}")

    with tempfile.TemporaryDirectory(prefix="llm-bench-recipes-results-") as tmp:
        extracted_root = extract_if_archive(args.input, Path(tmp))
        results_dir = find_results_dir(extracted_root)

        rows = load_summary_rows(results_dir)
        if not rows:
            sys.exit(f"summary.csv in {results_dir} has no rows")

        prov_path = results_dir / "provenance.json"
        prov = json.loads(prov_path.read_text()) if prov_path.is_file() else {}

        wb = Workbook()
        if args.split_by_tp:
            tp_groups = group_rows_by_tp(rows)
            first = True
            for tp, tp_rows in tp_groups.items():
                title = f"TP={tp}"
                ws = wb.active if first else wb.create_sheet(title)
                ws.title = title
                first = False
                build_results_sheet(ws, results_dir, tp_rows, include_tp_column=False)
        else:
            ws_results = wb.active
            ws_results.title = "Results"
            build_results_sheet(ws_results, results_dir, rows)

        ws_info = wb.create_sheet("Run Info")
        build_run_info_sheet(ws_info, prov)

        if prov.get("recipe_toml"):
            ws_recipe = wb.create_sheet("Recipe TOML")
            build_recipe_sheet(ws_recipe, prov)

        if args.output:
            out_path = args.output
        else:
            model = prov.get("model_name", "model")
            variant = prov.get("variant_name", "variant")
            ts = prov.get("timestamp_local", "results")
            suffix = "_by_tp" if args.split_by_tp else ""
            out_name = f"{model}_{variant}_{ts}_summary{suffix}.xlsx"
            base = args.input if args.input.is_dir() else args.input.parent
            out_path = base / out_name

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

        ok = sum(1 for r in rows if r.get("status") == "ok")
        print(f"Wrote {out_path}  ({ok}/{len(rows)} combos ok)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
